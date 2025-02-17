import pymysql, os, re
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# Load environment variables
load_dotenv()



# Change this if you need a different file name
output_file = "output.xlsx"
custom_date_format = "MM/DD/YYYY"






# Connect to MySQL
conn = pymysql.connect(
    host="db.conorbriggs.com.au",
    db='helios',
    user='helios-jamie',
    password='uyhgjikbed4r5896908ESRD!',
    port=3303,
    cursorclass=pymysql.cursors.DictCursor
)

# Query to fetch all data
query = """
    SELECT 
        datestamp as Date, 
        sheet_name, 
        market_cap_value as 'MCW', 
        equal_weight_value as 'EQW', 
        total_marketcap_value as 'TMC' 
    FROM 
        bf_raw_data
"""
cursor = conn.cursor()
cursor.execute(query)
data = cursor.fetchall()
df = pd.DataFrame(data)

# Convert datestamp to datetime 
df['Date'] = pd.to_datetime(df['Date'])

tab_renaming = {
    'Smart Contract Platform': "SCP",
    'Smart Contract Platform - Layer 0': "SCP - L0",
    'Smart Contract Platform - Layer 1': "SCP - L1",
    'Smart Contract Platform - Layer 2': "SCP - L2"
}

def rename_sheet(sheet_name):
    return tab_renaming.get(sheet_name, sheet_name)

df['sheet_name'] = df['sheet_name'].apply(rename_sheet)

sheet_order = [
    'top200', 'top200-altcoin', 
    'Digital Currency', 'SCP', 'DeFi', 'Infrastructure', 'Applications', 
    'Digital Currency - Memecoin', 'Digital Currency - Stablecoin', 'Digital Currency - Privacy', 'Digital Currency - General',
    'SCP - L0', 'SCP - L1', 'SCP - L2', 
    'DeFi - AMM', 'DeFi - Credit/Debit/Lending', 'DeFi - DAO', 'DeFi - DAO/Staking', 'DeFi - Derivatives & Tokenized Assets', 'DeFi - DEX', 'DeFi - Exchange (DEX)', 'DeFi - Lending', 'DeFi - Liquidity & Insurance', 'DeFi - Staking', 
    'Infrastructure - AI', 'Infrastructure - Computing & Networking', 'Infrastructure - IoT', 'Infrastructure - Oracle', 'Infrastructure - Storage/Data',  
    'Applications - Art, NFTs, & Collectibles',  'Applications - CEX',  'Applications - Exchange (CEX)',  'Applications - Gambling & Predictions',  'Applications - Gaming/Metaverse',  'Applications - Identity',  'Applications - Other',  'Applications - Social Network'  
]

def get_sheet_index(sheet_name):
    if sheet_name in sheet_order:
        return sheet_order.index(sheet_name)
    return len(sheet_order) # put at end if not found


existing_sheets = df['sheet_name'].unique()
sheets_sorted = sorted(existing_sheets, key=get_sheet_index)


with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
    workbook = writer.book
    date_format = workbook.add_format({'num_format': custom_date_format})
    
    for sheet in sheets_sorted:
        # Clean sheet name - replace invalid characters and truncate because excel is stupid
        sheet_name = re.sub(r'[\\/*?:"<>|]', '-', sheet)[:30]
        
        # Get data for this sheet
        group = df[df['sheet_name'] == sheet].copy()
        group = group.drop(columns=['sheet_name'])
        group = group.sort_values('Date')
        
        #renaming columns
        new_columns = ['Date'] + [f"{sheet} {col}" for col in group.columns if col != 'Date']
        group.columns = new_columns


        # Write to excel
        group.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get the worksheet
        worksheet = writer.sheets[sheet_name]
        
        # Format all cells in column A except the header (start from row 1, which is the second row in Excel)
        for row in range(1, len(group) + 1):
            worksheet.write(row, 0, group['Date'].iloc[row-1], date_format)


# Alt Season Index
cursor.execute("SELECT datestamp as `Date`, indicator_value as `Value` FROM bf_alt_season_indicator ORDER BY `Date` ASC")
data = cursor.fetchall()
df_alt_season_indi = pd.DataFrame(data)
df_alt_season_indi['Date'] = pd.to_datetime(df_alt_season_indi['Date'])

wb = load_workbook(output_file)
ws = wb.create_sheet(title="Alt Season Indicator", index=2)
ws.append(["Date", "Value"])

for row_idx, row in enumerate(df_alt_season_indi.itertuples(index=False), start=2):
    ws.append([row.Date, row.Value])
    ws[f"A{row_idx}"].number_format = custom_date_format

# wb.save(output_file)


# top 200 holdings (EQW)
cursor.execute("SELECT rebalance_date, coin_name, ticker, sector, subsector FROM `bf_index_holdings` WHERE index_name='top_200_eqw' ORDER BY rebalance_date DESC")
data = cursor.fetchall()
top_200_holdings_eqw = pd.DataFrame(data)
top_200_holdings_eqw['rebalance_date'] = pd.to_datetime(top_200_holdings_eqw['rebalance_date'])

ws = wb.create_sheet(title="Top 200 Holdings (EQW)", index=100)
ws.append(["Rebalance Date", 'Coin Name', 'ticker', 'sector', 'subsector'])

for row_idx, row in enumerate(top_200_holdings_eqw.itertuples(index=False), start=2):
    ws.append([row.rebalance_date, row.coin_name, row.ticker, row.sector, row.subsector])
    ws[f"A{row_idx}"].number_format = custom_date_format


# top 200 holdings (MC)
# cursor.execute("SELECT rebalance_date, coin_name, ticker, sector, subsector FROM `bf_index_holdings` WHERE index_name='top_200_mc' ORDER BY rebalance_date DESC")
# data = cursor.fetchall()
# top_200_holdings_eqw = pd.DataFrame(data)
# top_200_holdings_eqw['rebalance_date'] = pd.to_datetime(top_200_holdings_eqw['rebalance_date'])

# ws = wb.create_sheet(title="Top 200 Holdings (MC)", index=100)
# ws.append(["Rebalance Date", 'Coin Name', 'ticker', 'sector', 'subsector'])

# for row_idx, row in enumerate(top_200_holdings_eqw.itertuples(index=False), start=2):
#     ws.append([row.rebalance_date, row.coin_name, row.ticker, row.sector, row.subsector])
#     ws[f"A{row_idx}"].number_format = custom_date_format






wb.save(output_file)

cursor.close()
conn.close()

print(f"Excel file '{output_file}' created successfully!")
